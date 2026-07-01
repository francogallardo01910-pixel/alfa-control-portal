import streamlit as st
from supabase import create_client
import pandas as pd
import hashlib
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.page import PageMargins

# =========================================================
# CONFIGURACIÓN GENERAL
# =========================================================

st.set_page_config(
    page_title="ALFA Control Portal v2.0",
    page_icon="📦",
    layout="wide"
)

SUPABASE_URL = "https://dpejdcfmovrjvshksdlq.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImRwZWpkY2Ztb3ZyanZzaGtzZGxxIiwicm9sZSI6ImFub24iLCJpYXQiOjE3ODI3NjUxMTQsImV4cCI6MjA5ODM0MTExNH0.LQJEGDUpyZZ4Yw7P0s0XFhlL6OzHOXfygJuOSSG1MtM"

supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

CLAVE_ADMIN = "1234"
DIA_CIERRE_SOLICITUD = 10
DIA_RETIRO = 15
MONTO_MAXIMO_SOLICITUD = 120000

PRECIOS_ALIMENTO_DEFAULT = {
    "Sanson 25 kg": 19000,
    "Sanson 9kg": 7000,
    "Alfa Tradicional Adulto 25 kg": 22000,
    "Alfa Tradicional Cachorro 18 kg": 18000,
    "Felino 20 kg": 19000,
    "Alfa Premium Adulto 20 kg": 21000,
    "Alfa Premium Senior 20 kg": 22000,
    "Alfa Premium Ad Raza Pequeña 10 kg": 11000,
    "Alfa Premium Sen Raza Pequeña 10 kg": 12000,
    "Alfa Premium Cachorro 10 kg": 12000,
    "Alfa Premium Cach Raza Pequeña 10 kg": 12000,
    "Alfa Cat 10 kg": 15000,
    "ALFA PREMIUM CAT 2 KG": 6000,
    "Alfa Premium Adulto 3 kg": 7000,
    "Alfa Premium Senior 3 kg": 7000,
    "Alfa Premium Ad Raza Pequeña 3 kg": 7000,
    "Alfa Premium Cachorro 3 kg": 7000,
    "Alfa Premium Sen Raza Pequeña 3 kg": 7000,
    "Alfa Premium Cach Raza Pequeña 3 kg": 7000,
}

# =========================================================
# DISEÑO
# =========================================================

st.markdown("""
<style>
.main { background-color: #f5f7f5; }
.block-container { padding-top: 1.5rem; }
.header {
    background: linear-gradient(90deg, #0b6b3a, #159957);
    padding: 26px;
    border-radius: 18px;
    color: white;
    margin-bottom: 24px;
}
.metric-card {
    background: white;
    padding: 18px;
    border-radius: 14px;
    text-align: center;
    box-shadow: 0px 3px 10px rgba(0,0,0,0.08);
    min-height: 110px;
}
.big-number {
    font-size: 30px;
    font-weight: bold;
    color: #0b6b3a;
}
.row-card {
    background: white;
    padding: 14px;
    border-radius: 12px;
    box-shadow: 0px 2px 8px rgba(0,0,0,0.06);
    margin-bottom: 8px;
}
.info-card {
    background: white;
    border-left: 6px solid #0b6b3a;
    padding: 16px;
    border-radius: 12px;
    margin-bottom: 10px;
}
</style>
""", unsafe_allow_html=True)

# =========================================================
# FUNCIONES BASE
# =========================================================

def limpiar_texto(valor):
    if pd.isna(valor):
        return ""
    return str(valor).strip()


def normalizar_nombre(valor):
    return limpiar_texto(valor).upper()


def formato_pesos(valor):
    try:
        valor = int(valor)
        if valor == 0:
            return "$-"
        return "$" + f"{valor:,}".replace(",", ".")
    except Exception:
        return "$-"


def convertir_excel(df, nombre_hoja="Datos"):
    salida = BytesIO()
    with pd.ExcelWriter(salida, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=nombre_hoja[:31])
    return salida.getvalue()


def mes_actual_inicio():
    hoy = datetime.now()
    return f"{hoy.year}-{hoy.month:02d}-01"


def solicitudes_abiertas():
    return datetime.now().day <= DIA_CIERRE_SOLICITUD


def validar_admin(clave):
    # Mantiene compatibilidad con módulos antiguos, pero si ya inició sesión
    # como administrador no vuelve a pedir la clave simple.
    if st.session_state.get("admin_logueado"):
        return True
    return clave == CLAVE_ADMIN


def encriptar_clave(clave):
    return hashlib.sha256(str(clave).encode("utf-8")).hexdigest()


def obtener_usuarios_admin_df():
    try:
        res = supabase.table("usuarios_admin").select("*").order("usuario").execute()
        return pd.DataFrame(res.data)
    except Exception:
        return pd.DataFrame()


def validar_credenciales_admin(usuario, clave):
    usuario = limpiar_texto(usuario).lower()
    clave = limpiar_texto(clave)

    if not usuario or not clave:
        return False, "Debe ingresar usuario y clave."

    try:
        res = (
            supabase.table("usuarios_admin")
            .select("*")
            .eq("usuario", usuario)
            .eq("activo", True)
            .execute()
        )
        if res.data:
            admin = res.data[0]
            if admin.get("clave_hash") == encriptar_clave(clave):
                return True, admin.get("nombre") or usuario
            return False, "Usuario o clave incorrecta."
    except Exception:
        pass

    # Usuario de respaldo inicial. Sirve para entrar la primera vez y crear usuarios.
    if usuario == "franco" and clave == CLAVE_ADMIN:
        return True, "franco"

    return False, "Usuario o clave incorrecta."


def login_admin():
    st.markdown("""
    <div class="header">
        <h1>🔐 Acceso administrador</h1>
        <h3>Ingrese usuario y clave para administrar el portal</h3>
    </div>
    """, unsafe_allow_html=True)

    c1, c2 = st.columns(2)
    usuario = c1.text_input("Usuario")
    clave = c2.text_input("Clave", type="password")

    if st.button("Ingresar", use_container_width=True):
        ok, nombre = validar_credenciales_admin(usuario, clave)
        if ok:
            st.session_state["admin_logueado"] = True
            st.session_state["usuario_admin"] = limpiar_texto(usuario).lower()
            st.session_state["nombre_admin"] = nombre
            st.success("Acceso correcto.")
            st.rerun()
        else:
            st.error(nombre)

    st.info("Usuario inicial de respaldo: franco / 1234. Después puede crear usuarios reales en el módulo Usuarios admin.")


def agregar_usuario_admin(usuario, nombre, clave, activo=True):
    usuario = limpiar_texto(usuario).lower()
    nombre = limpiar_texto(nombre) or usuario

    if not usuario or not clave:
        return False, "Debe ingresar usuario y clave."

    if len(clave) < 4:
        return False, "La clave debe tener al menos 4 caracteres."

    try:
        existe = supabase.table("usuarios_admin").select("*").eq("usuario", usuario).execute()
        if existe.data:
            return False, "Ese usuario ya existe."

        supabase.table("usuarios_admin").insert({
            "usuario": usuario,
            "nombre": nombre,
            "clave_hash": encriptar_clave(clave),
            "activo": bool(activo),
            "creado": datetime.now().isoformat()
        }).execute()
        return True, "Usuario administrador creado correctamente."
    except Exception as e:
        return False, f"No se pudo crear usuario. Revise que la tabla usuarios_admin exista en Supabase. Error: {e}"


def actualizar_usuario_admin(id_usuario, nombre, nueva_clave=None, activo=True):
    data = {
        "nombre": limpiar_texto(nombre),
        "activo": bool(activo)
    }
    if nueva_clave:
        if len(nueva_clave) < 4:
            return False, "La nueva clave debe tener al menos 4 caracteres."
        data["clave_hash"] = encriptar_clave(nueva_clave)

    try:
        supabase.table("usuarios_admin").update(data).eq("id", int(id_usuario)).execute()
        return True, "Usuario actualizado correctamente."
    except Exception as e:
        return False, f"No se pudo actualizar usuario: {e}"


def eliminar_usuario_admin(id_usuario):
    try:
        supabase.table("usuarios_admin").delete().eq("id", int(id_usuario)).execute()
        return True, "Usuario eliminado correctamente."
    except Exception as e:
        return False, f"No se pudo eliminar usuario: {e}"


# =========================================================
# SUPABASE: TRABAJADORES
# =========================================================

def obtener_trabajadores_df():
    try:
        res = supabase.table("trabajadores").select("*").order("nombre").execute()
        df = pd.DataFrame(res.data)
        if not df.empty:
            for col in ["area", "cargo"]:
                if col not in df.columns:
                    df[col] = ""
            if "activo" not in df.columns:
                df["activo"] = True
        return df
    except Exception as e:
        st.error(f"No se pudo cargar trabajadores: {e}")
        return pd.DataFrame()


def obtener_trabajadores_dict():
    df = obtener_trabajadores_df()
    if df.empty:
        return {}
    return {r["rut"]: r.get("nombre", "") for _, r in df.iterrows()}


def agregar_trabajador(datos):
    rut = limpiar_texto(datos.get("rut"))
    nombre = normalizar_nombre(datos.get("nombre"))

    if not rut or not nombre:
        return False, "Debe ingresar RUT y nombre."

    existe = supabase.table("trabajadores").select("*").eq("rut", rut).execute()
    if existe.data:
        return False, "Este RUT ya existe."

    data = {
        "rut": rut,
        "nombre": nombre,
        "area": normalizar_nombre(datos.get("area")),
        "cargo": normalizar_nombre(datos.get("cargo")),
        "activo": bool(datos.get("activo", True)),
    }

    try:
        supabase.table("trabajadores").insert(data).execute()
    except Exception:
        supabase.table("trabajadores").insert({"rut": rut, "nombre": nombre}).execute()

    registrar_historial("AGREGAR", rut, nombre, "Trabajador agregado")
    return True, "Trabajador agregado correctamente."


def actualizar_trabajador(rut_original, datos):
    rut_nuevo = limpiar_texto(datos.get("rut"))
    nombre_nuevo = normalizar_nombre(datos.get("nombre"))

    if not rut_nuevo or not nombre_nuevo:
        return False, "Debe ingresar RUT y nombre."

    if rut_nuevo != rut_original:
        existe = supabase.table("trabajadores").select("*").eq("rut", rut_nuevo).execute()
        if existe.data:
            return False, "El nuevo RUT ya existe en otro trabajador."

    data = {
        "rut": rut_nuevo,
        "nombre": nombre_nuevo,
        "area": normalizar_nombre(datos.get("area")),
        "cargo": normalizar_nombre(datos.get("cargo")),
        "activo": bool(datos.get("activo", True)),
    }

    try:
        supabase.table("trabajadores").update(data).eq("rut", rut_original).execute()
    except Exception:
        supabase.table("trabajadores").update({"rut": rut_nuevo, "nombre": nombre_nuevo}).eq("rut", rut_original).execute()

    registrar_historial("EDITAR", rut_nuevo, nombre_nuevo, f"RUT anterior: {rut_original}")
    return True, "Trabajador actualizado correctamente."


def eliminar_trabajador(rut, nombre=""):
    supabase.table("trabajadores").delete().eq("rut", rut).execute()
    registrar_historial("ELIMINAR", rut, nombre, "Trabajador eliminado")
    return True, "Trabajador eliminado correctamente."


def registrar_historial(accion, rut="", nombre="", detalle=""):
    try:
        supabase.table("historial_trabajadores").insert({
            "fecha": datetime.now().isoformat(),
            "accion": accion,
            "rut": rut,
            "nombre": nombre,
            "detalle": detalle
        }).execute()
    except Exception:
        pass


def cargar_historial_df():
    try:
        res = supabase.table("historial_trabajadores").select("*").order("fecha", desc=True).execute()
        return pd.DataFrame(res.data)
    except Exception:
        return pd.DataFrame()


def importar_trabajadores_desde_excel(archivo):
    df = pd.read_excel(archivo)
    df.columns = [str(c).strip().lower() for c in df.columns]

    if "rut" not in df.columns or "nombre" not in df.columns:
        return False, "El Excel debe tener al menos las columnas rut y nombre.", pd.DataFrame()

    for col in ["area", "cargo"]:
        if col not in df.columns:
            df[col] = ""
    if "activo" not in df.columns:
        df["activo"] = True

    agregados = 0
    actualizados = 0
    errores = []

    for idx, fila in df.iterrows():
        rut = limpiar_texto(fila.get("rut"))
        nombre = normalizar_nombre(fila.get("nombre"))

        if not rut or not nombre:
            errores.append(f"Fila {idx + 2}: falta RUT o nombre.")
            continue

        activo = fila.get("activo", True)
        if isinstance(activo, str):
            activo = activo.lower().strip() in ["si", "sí", "true", "1", "activo"]

        datos = {
            "rut": rut,
            "nombre": nombre,
            "area": fila.get("area", ""),
            "cargo": fila.get("cargo", ""),
            "activo": bool(activo),
        }

        try:
            existe = supabase.table("trabajadores").select("*").eq("rut", rut).execute()
            if existe.data:
                actualizar_trabajador(rut, datos)
                actualizados += 1
            else:
                agregar_trabajador(datos)
                agregados += 1
        except Exception as e:
            errores.append(f"Fila {idx + 2}: {e}")

    msg = f"Importación terminada. Agregados: {agregados}. Actualizados: {actualizados}. Errores: {len(errores)}."
    return True, msg, pd.DataFrame({"Errores": errores})


# =========================================================
# SUPABASE: PRODUCTOS
# =========================================================

def obtener_productos_df():
    try:
        res = supabase.table("productos_alimento").select("*").order("producto").execute()
        df = pd.DataFrame(res.data)
        if not df.empty:
            if "activo" not in df.columns:
                df["activo"] = True
            if "precio" not in df.columns:
                df["precio"] = 0
        return df
    except Exception as e:
        st.error(f"No se pudo cargar productos_alimento: {e}")
        return pd.DataFrame()


def obtener_productos_activos():
    try:
        res = supabase.table("productos_alimento").select("*").eq("activo", True).order("producto").execute()
        productos = [p["producto"] for p in res.data]
        return productos
    except Exception:
        return list(PRECIOS_ALIMENTO_DEFAULT.keys())


def agregar_producto_alimento(producto, precio=0, activo=True):
    producto = limpiar_texto(producto)
    if not producto:
        return False, "Debe ingresar producto."

    existe = supabase.table("productos_alimento").select("*").eq("producto", producto).execute()
    if existe.data:
        return False, "Este producto ya existe."

    try:
        supabase.table("productos_alimento").insert({
            "producto": producto,
            "precio": int(precio or 0),
            "activo": bool(activo)
        }).execute()
    except Exception as e:
        return False, f"No se pudo agregar producto: {e}"

    return True, "Producto agregado correctamente."


def actualizar_producto_alimento(id_producto, producto, precio=0, activo=True):
    producto = limpiar_texto(producto)
    if not producto:
        return False, "Debe ingresar producto."

    try:
        supabase.table("productos_alimento").update({
            "producto": producto,
            "precio": int(precio or 0),
            "activo": bool(activo)
        }).eq("id", int(id_producto)).execute()
    except Exception as e:
        return False, f"No se pudo actualizar producto: {e}"

    return True, "Producto actualizado correctamente."


def eliminar_producto_alimento(id_producto):
    try:
        supabase.table("productos_alimento").delete().eq("id", int(id_producto)).execute()
        return True, "Producto eliminado correctamente."
    except Exception as e:
        return False, f"No se pudo eliminar producto: {e}"


def cargar_precios_default_supabase():
    creados = 0
    actualizados = 0
    errores = []

    for producto, precio in PRECIOS_ALIMENTO_DEFAULT.items():
        try:
            existe = supabase.table("productos_alimento").select("*").eq("producto", producto).execute()
            if existe.data:
                supabase.table("productos_alimento").update({
                    "precio": int(precio),
                    "activo": True
                }).eq("id", int(existe.data[0]["id"])).execute()
                actualizados += 1
            else:
                supabase.table("productos_alimento").insert({
                    "producto": producto,
                    "precio": int(precio),
                    "activo": True
                }).execute()
                creados += 1
        except Exception as e:
            errores.append(f"{producto}: {e}")

    return creados, actualizados, errores


def obtener_precios_productos():
    """
    Usa primero los precios oficiales cargados en el código.
    Si Supabase tiene precio mayor a 0, lo reemplaza.
    Si Supabase tiene precio 0, vacío o nulo, mantiene el precio oficial.
    """
    precios = PRECIOS_ALIMENTO_DEFAULT.copy()

    try:
        res = supabase.table("productos_alimento").select("producto, precio").execute()

        for p in res.data:
            producto = p.get("producto")
            precio_supabase = int(p.get("precio") or 0)

            if producto:
                # Solo reemplaza si Supabase tiene un precio real mayor a 0
                if precio_supabase > 0:
                    precios[producto] = precio_supabase
                else:
                    # Si está en 0, intenta mantener precio oficial por nombre normalizado
                    producto_norm = normalizar_producto_para_precio(producto)
                    for prod_oficial, precio_oficial in PRECIOS_ALIMENTO_DEFAULT.items():
                        if normalizar_producto_para_precio(prod_oficial) == producto_norm:
                            precios[producto] = precio_oficial
                            break

    except Exception:
        pass

    return precios


def normalizar_producto_para_precio(texto):
    """
    Normaliza nombres de productos para encontrar precios aunque tengan
    mayúsculas, espacios dobles o variaciones menores.
    """
    texto = str(texto or "").strip().lower()
    texto = texto.replace(".", "")
    texto = texto.replace("  ", " ")
    texto = " ".join(texto.split())
    texto = texto.replace("sanson 9 kg", "sanson 9kg")
    texto = texto.replace("sanson 9k", "sanson 9kg")
    return texto


def buscar_precio_producto(nombre_producto, precios):
    if not nombre_producto:
        return 0

    # Búsqueda exacta
    if nombre_producto in precios:
        return int(precios.get(nombre_producto, 0) or 0)

    # Búsqueda normalizada
    nombre_limpio = normalizar_producto_para_precio(nombre_producto)

    for prod, precio in precios.items():
        prod_limpio = normalizar_producto_para_precio(prod)
        if prod_limpio == nombre_limpio:
            return int(precio or 0)

    # Búsqueda parcial controlada
    for prod, precio in precios.items():
        prod_limpio = normalizar_producto_para_precio(prod)
        if nombre_limpio in prod_limpio or prod_limpio in nombre_limpio:
            return int(precio or 0)

    return 0


# =========================================================
# SOLICITUDES Y REPORTES
# =========================================================

def obtener_solicitudes_df():
    try:
        res = supabase.table("solicitudes_alimento").select("*").order("fecha", desc=True).execute()
        df = pd.DataFrame(res.data)
        return df
    except Exception as e:
        st.error(f"No se pudo cargar solicitudes: {e}")
        return pd.DataFrame()


def preparar_solicitudes_con_nombre():
    df = obtener_solicitudes_df()
    if df.empty:
        return df

    nombres = obtener_trabajadores_dict()
    df["nombre"] = df["rut"].map(nombres).fillna("Sin nombre")
    df["fecha_dt"] = pd.to_datetime(df["fecha"], errors="coerce")
    df["mes"] = df["fecha_dt"].dt.strftime("%Y-%m")

    if "monto_solicitado" not in df.columns:
        df["monto_solicitado"] = 0

    return df


def convertir_excel_formato_quincenal(df_filtrado, mes_seleccionado):
    precios = obtener_precios_productos()

    wb = Workbook()
    ws = wb.active
    ws.title = "Alimento y anticipo"

    verde = "0B6B3A"
    verde_medio = "159957"
    verde_claro = "D9EAD3"
    gris = "F2F2F2"
    gris_oscuro = "D9D9D9"
    blanco = "FFFFFF"
    borde_fino = Side(style="thin", color="B7B7B7")
    borde = Border(left=borde_fino, right=borde_fino, top=borde_fino, bottom=borde_fino)

    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.45, bottom=0.45)

    ws.merge_cells("A1:I1")
    ws["A1"] = f"ALFA CONTROL PORTAL - SOLICITUD QUINCENAL {mes_seleccionado}"
    ws["A1"].font = Font(bold=True, size=16, color=blanco)
    ws["A1"].fill = PatternFill("solid", fgColor=verde)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:E2")
    ws["A2"] = "ALIMENTO DE MASCOTA QUINCENAL"
    ws["A2"].font = Font(bold=True, size=13, color=blanco)
    ws["A2"].fill = PatternFill("solid", fgColor=verde_medio)
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("G2:I2")
    ws["G2"] = "ANTICIPO MONETARIO"
    ws["G2"].font = Font(bold=True, size=13, color=blanco)
    ws["G2"].fill = PatternFill("solid", fgColor=verde_medio)
    ws["G2"].alignment = Alignment(horizontal="center")

    for col, titulo in enumerate(["Nombre", "Producto", "PRECIO ALIMENTO", "Cant.", "TOTAL"], start=1):
        celda = ws.cell(row=3, column=col, value=titulo)
        celda.font = Font(bold=True)
        celda.fill = PatternFill("solid", fgColor=verde_claro)
        celda.border = borde
        celda.alignment = Alignment(horizontal="center", vertical="center")

    for col, titulo in enumerate(["Nombre completo", "RUT", "Anticipo Monetario"], start=7):
        celda = ws.cell(row=3, column=col, value=titulo)
        celda.font = Font(bold=True)
        celda.fill = PatternFill("solid", fgColor=verde_claro)
        celda.border = borde
        celda.alignment = Alignment(horizontal="center", vertical="center")

    fila_excel = 4
    total_cantidad = 0
    total_alimento = 0

    for _, fila in df_filtrado.iterrows():
        nombre = fila.get("nombre", "")
        for i in [1, 2, 3]:
            producto = fila.get(f"producto{i}")
            cantidad = fila.get(f"cantidad{i}")

            if producto and pd.notna(cantidad):
                cantidad = int(cantidad)
                precio = buscar_precio_producto(producto, precios)
                total = precio * cantidad
                valores = [nombre, producto, precio, cantidad, total]

                for col, valor in enumerate(valores, start=1):
                    celda = ws.cell(row=fila_excel, column=col, value=valor)
                    celda.border = borde
                    celda.alignment = Alignment(vertical="center", wrap_text=True)

                ws.cell(row=fila_excel, column=3).number_format = '$#,##0;-$#,##0;"$-"'
                ws.cell(row=fila_excel, column=5).number_format = '$#,##0;-$#,##0;"$-"'

                total_cantidad += cantidad
                total_alimento += total
                fila_excel += 1

    fila_total_alimento = fila_excel
    ws.cell(row=fila_total_alimento, column=2, value="TOTAL")
    ws.cell(row=fila_total_alimento, column=3, value="TOTAL ALIMENTO")
    ws.cell(row=fila_total_alimento, column=4, value=total_cantidad)
    ws.cell(row=fila_total_alimento, column=5, value=total_alimento)
    ws.cell(row=fila_total_alimento, column=5).number_format = '$#,##0;-$#,##0;"$-"'

    for col in range(1, 6):
        celda = ws.cell(row=fila_total_alimento, column=col)
        celda.font = Font(bold=True)
        celda.fill = PatternFill("solid", fgColor=gris_oscuro)
        celda.border = borde
        celda.alignment = Alignment(horizontal="center", vertical="center")

    fila_anticipo = 4
    total_anticipo = 0

    for _, fila in df_filtrado.iterrows():
        monto = int(fila.get("monto_solicitado", 0)) if pd.notna(fila.get("monto_solicitado", 0)) else 0
        total_anticipo += monto
        valores = [fila.get("nombre", ""), fila.get("rut", ""), monto]

        for col, valor in zip(range(7, 10), valores):
            celda = ws.cell(row=fila_anticipo, column=col, value=valor)
            celda.border = borde
            celda.alignment = Alignment(vertical="center", wrap_text=True)

        ws.cell(row=fila_anticipo, column=9).number_format = '$#,##0;-$#,##0;"$-"'
        fila_anticipo += 1

    ws.cell(row=fila_anticipo, column=8, value="TOTAL ANTICIPO")
    ws.cell(row=fila_anticipo, column=9, value=total_anticipo)
    ws.cell(row=fila_anticipo, column=9).number_format = '$#,##0;-$#,##0;"$-"'

    for col in range(7, 10):
        celda = ws.cell(row=fila_anticipo, column=col)
        celda.font = Font(bold=True)
        celda.fill = PatternFill("solid", fgColor=gris_oscuro)
        celda.border = borde
        celda.alignment = Alignment(horizontal="center", vertical="center")

    fila_resumen = max(fila_total_alimento, fila_anticipo) + 2
    ws.merge_cells(start_row=fila_resumen, start_column=1, end_row=fila_resumen, end_column=9)
    ws.cell(row=fila_resumen, column=1, value="RESUMEN GENERAL")
    ws.cell(row=fila_resumen, column=1).font = Font(bold=True, size=12, color=blanco)
    ws.cell(row=fila_resumen, column=1).fill = PatternFill("solid", fgColor=verde)
    ws.cell(row=fila_resumen, column=1).alignment = Alignment(horizontal="center")

    resumen_datos = [
        ("Total sacos alimento", total_cantidad),
        ("Total valorizado alimento", total_alimento),
        ("Total anticipo monetario", total_anticipo),
        ("Total general alimento + anticipo", total_alimento + total_anticipo),
    ]

    for idx, (concepto, valor) in enumerate(resumen_datos, start=fila_resumen + 1):
        ws.cell(row=idx, column=1, value=concepto)
        ws.cell(row=idx, column=2, value=valor)
        ws.cell(row=idx, column=1).font = Font(bold=True)
        ws.cell(row=idx, column=1).fill = PatternFill("solid", fgColor=gris)
        ws.cell(row=idx, column=2).fill = PatternFill("solid", fgColor=gris)
        ws.cell(row=idx, column=1).border = borde
        ws.cell(row=idx, column=2).border = borde

        if "sacos" not in concepto.lower():
            ws.cell(row=idx, column=2).number_format = '$#,##0;-$#,##0;"$-"'

    for col, ancho in {"A": 30, "B": 38, "C": 18, "D": 10, "E": 18, "F": 4, "G": 32, "H": 18, "I": 20}.items():
        ws.column_dimensions[col].width = ancho

    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 22

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws.freeze_panes = "A4"

    salida = BytesIO()
    wb.save(salida)
    return salida.getvalue()




# =========================================================
# SUPABASE: RESET Y LIMPIEZA DE PRUEBAS
# =========================================================

def eliminar_todos_registros_tabla(tabla):
    """
    Elimina todos los registros de una tabla usando el campo id.
    Mantiene la estructura de Supabase y solo borra datos.
    """
    try:
        supabase.table(tabla).delete().neq("id", 0).execute()
        return True, f"Registros eliminados de {tabla}."
    except Exception as e:
        return False, f"No se pudo limpiar {tabla}: {e}"


def eliminar_solicitud_por_id(id_solicitud):
    try:
        supabase.table("solicitudes_alimento").delete().eq("id", int(id_solicitud)).execute()
        return True, f"Solicitud ID {id_solicitud} eliminada correctamente."
    except Exception as e:
        return False, f"No se pudo eliminar la solicitud: {e}"


def marcar_solicitudes_pendientes():
    try:
        supabase.table("solicitudes_alimento").update({"estado": "Pendiente"}).neq("id", 0).execute()
        return True, "Todas las solicitudes fueron marcadas como Pendiente."
    except Exception as e:
        return False, f"No se pudo actualizar solicitudes: {e}"


# =========================================================
# UI: MENÚ Y ACCESO
# =========================================================

# Enlace normal / QR: abre directo la solicitud del trabajador y oculta menú.
# Enlace administrador: agregue ?admin=1 al final del enlace público.
modo_admin = "admin" in st.query_params

if not modo_admin:
    st.markdown("""
    <style>
        [data-testid="stSidebar"] {display: none;}
        .block-container {max-width: 950px;}
    </style>
    """, unsafe_allow_html=True)
    menu = "Solicitud trabajador"
else:
    if "admin_logueado" not in st.session_state:
        st.session_state["admin_logueado"] = False

    if not st.session_state["admin_logueado"]:
        login_admin()
        st.stop()

    st.sidebar.title("📦 ALFA Control")
    st.sidebar.success(f"Admin: {st.session_state.get('nombre_admin', st.session_state.get('usuario_admin', ''))}")

    if st.sidebar.button("Cerrar sesión"):
        st.session_state["admin_logueado"] = False
        st.session_state["usuario_admin"] = ""
        st.session_state["nombre_admin"] = ""
        st.rerun()

    menu = st.sidebar.radio(
        "Seleccione módulo",
        [
            "Inicio",
            "Solicitud trabajador",
            "Administrador",
            "Trabajadores",
            "Productos",
            "Usuarios admin",
            "Reportes",
            "Reset / limpieza",
            "Configuración"
        ]
    )

# =========================================================
# MÓDULO: INICIO
# =========================================================

if menu == "Inicio":
    st.markdown("""
    <div class="header">
        <h1>📦 ALFA Control Portal v2.0</h1>
        <h3>Solicitud quincenal de alimento y anticipo monetario</h3>
    </div>
    """, unsafe_allow_html=True)

    df = preparar_solicitudes_con_nombre()
    df_trab = obtener_trabajadores_df()
    df_prod = obtener_productos_df()

    total_solicitudes = len(df) if not df.empty else 0
    total_trabajadores = len(df_trab) if not df_trab.empty else 0
    total_productos = len(df_prod) if not df_prod.empty else 0
    total_monto = int(df["monto_solicitado"].fillna(0).sum()) if not df.empty and "monto_solicitado" in df.columns else 0

    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.markdown(f'<div class="metric-card"><p>Solicitudes</p><div class="big-number">{total_solicitudes}</div></div>', unsafe_allow_html=True)
    with col2:
        st.markdown(f'<div class="metric-card"><p>Trabajadores</p><div class="big-number">{total_trabajadores}</div></div>', unsafe_allow_html=True)
    with col3:
        st.markdown(f'<div class="metric-card"><p>Productos</p><div class="big-number">{total_productos}</div></div>', unsafe_allow_html=True)
    with col4:
        st.markdown(f'<div class="metric-card"><p>Anticipo total</p><div class="big-number">{formato_pesos(total_monto)}</div></div>', unsafe_allow_html=True)

    st.info(f"Solicitudes abiertas hasta el día {DIA_CIERRE_SOLICITUD}. Retiro programado para el día {DIA_RETIRO}.")


# =========================================================
# MÓDULO: SOLICITUD TRABAJADOR
# =========================================================

if menu == "Solicitud trabajador":
    st.markdown("""
    <div class="header">
        <h1>👤 Solicitud trabajador</h1>
        <h3>Alimento de mascota y anticipo monetario</h3>
    </div>
    """, unsafe_allow_html=True)

    if solicitudes_abiertas():
        st.success(f"Solicitudes abiertas hasta el día {DIA_CIERRE_SOLICITUD}. Retiro programado para el día {DIA_RETIRO}.")
    else:
        st.error(f"Las solicitudes de este mes ya fueron cerradas. Se reciben hasta el día {DIA_CIERRE_SOLICITUD}.")
        st.stop()

    rut = st.text_input("Ingrese su RUT", placeholder="Ej: 12345678-9").strip()

    if rut:
        res = supabase.table("trabajadores").select("*").eq("rut", rut).execute()

        if not res.data:
            st.error("Trabajador no encontrado.")
            st.stop()

        trabajador = res.data[0]
        if trabajador.get("activo") is False:
            st.error("Este trabajador está inactivo.")
            st.stop()

        st.success(f"Bienvenido/a {trabajador.get('nombre', '')}")

        solicitud_existente = (
            supabase.table("solicitudes_alimento")
            .select("*")
            .eq("rut", rut)
            .gte("fecha", mes_actual_inicio())
            .execute()
        )

        if solicitud_existente.data:
            st.warning("Usted ya realizó una solicitud este mes.")
            st.dataframe(pd.DataFrame(solicitud_existente.data), use_container_width=True)
            st.stop()

        productos = obtener_productos_activos()
        if not productos:
            st.error("No hay productos activos disponibles.")
            st.stop()

        st.markdown("### 👤 Datos del trabajador")
        c1, c2, c3 = st.columns(3)
        c1.info(f"**Nombre:** {trabajador.get('nombre', '')}")
        c2.info(f"**Área:** {trabajador.get('area', 'SIN ÁREA')}")
        c3.info(f"**Cargo:** {trabajador.get('cargo', 'SIN CARGO')}")

        st.markdown("### 💰 Anticipo monetario")
        monto_solicitado = st.number_input(
            "Monto solicitado",
            min_value=0,
            max_value=MONTO_MAXIMO_SOLICITUD,
            step=1000,
            help="Máximo permitido $120.000"
        )

        porcentaje = monto_solicitado / MONTO_MAXIMO_SOLICITUD if MONTO_MAXIMO_SOLICITUD else 0
        st.progress(porcentaje)
        st.info(f"Límite máximo: {formato_pesos(MONTO_MAXIMO_SOLICITUD)}")

        st.markdown("### 🍖 Alimentos que necesita")

        producto1 = st.selectbox("Alimento 1", productos)
        cantidad1 = st.number_input("Cantidad alimento 1", min_value=1, max_value=3, step=1)

        agregar2 = st.checkbox("➕ Agregar alimento 2")
        producto2, cantidad2 = None, None
        if agregar2:
            producto2 = st.selectbox("Alimento 2", productos, key="producto2")
            cantidad2 = st.number_input("Cantidad alimento 2", min_value=1, max_value=3, step=1, key="cantidad2")

        agregar3 = st.checkbox("➕ Agregar alimento 3")
        producto3, cantidad3 = None, None
        if agregar3:
            producto3 = st.selectbox("Alimento 3", productos, key="producto3")
            cantidad3 = st.number_input("Cantidad alimento 3", min_value=1, max_value=3, step=1, key="cantidad3")

        resumen = [{"Alimento": producto1, "Cantidad": int(cantidad1)}]
        if agregar2 and producto2:
            resumen.append({"Alimento": producto2, "Cantidad": int(cantidad2)})
        if agregar3 and producto3:
            resumen.append({"Alimento": producto3, "Cantidad": int(cantidad3)})

        st.markdown("### 📋 Resumen antes de enviar")
        st.dataframe(pd.DataFrame(resumen), use_container_width=True)
        st.info(f"Anticipo solicitado: {formato_pesos(monto_solicitado)}")

        if st.button("✅ Enviar solicitud", use_container_width=True):
            if monto_solicitado < 0 or monto_solicitado > MONTO_MAXIMO_SOLICITUD:
                st.error("El monto solicitado no puede superar $120.000.")
                st.stop()

            supabase.table("solicitudes_alimento").insert({
                "rut": rut,
                "monto_solicitado": int(monto_solicitado),
                "producto1": producto1,
                "cantidad1": int(cantidad1),
                "producto2": producto2,
                "cantidad2": int(cantidad2) if cantidad2 else None,
                "producto3": producto3,
                "cantidad3": int(cantidad3) if cantidad3 else None,
                "estado": "Pendiente"
            }).execute()

            st.success("Solicitud enviada correctamente.")
            st.rerun()


# =========================================================
# MÓDULO: ADMINISTRADOR
# =========================================================

if menu == "Administrador":
    st.markdown("""
    <div class="header">
        <h1>👨‍💼 Panel administrador</h1>
        <h3>Control de solicitudes, entregas y Excel quincenal</h3>
    </div>
    """, unsafe_allow_html=True)

    df = preparar_solicitudes_con_nombre()
    if df.empty:
        st.info("No hay solicitudes registradas.")
        st.stop()

    meses = sorted(df["mes"].dropna().unique(), reverse=True)
    mes_seleccionado = st.selectbox("Seleccione mes", meses)

    df = df[df["mes"] == mes_seleccionado].copy()

    filtro_estado = st.selectbox("Estado", ["Todos", "Pendiente", "Entregado"])
    buscar = st.text_input("Buscar por RUT o nombre")

    df_filtrado = df.copy()
    if filtro_estado != "Todos":
        df_filtrado = df_filtrado[df_filtrado["estado"] == filtro_estado]
    if buscar:
        texto = buscar.lower()
        df_filtrado = df_filtrado[
            df_filtrado["rut"].astype(str).str.lower().str.contains(texto, na=False) |
            df_filtrado["nombre"].astype(str).str.lower().str.contains(texto, na=False)
        ]

    total_solicitudes = len(df_filtrado)
    pendientes = len(df_filtrado[df_filtrado["estado"] == "Pendiente"])
    entregadas = len(df_filtrado[df_filtrado["estado"] == "Entregado"])
    total_monto = int(df_filtrado["monto_solicitado"].fillna(0).sum()) if "monto_solicitado" in df_filtrado.columns else 0

    total_sacos = 0
    resumen_productos = {}
    precios = obtener_precios_productos()
    total_valorizado_alimento = 0

    for _, fila in df_filtrado.iterrows():
        for i in [1, 2, 3]:
            producto = fila.get(f"producto{i}")
            cantidad = fila.get(f"cantidad{i}")
            if producto and pd.notna(cantidad):
                cantidad = int(cantidad)
                total_sacos += cantidad
                resumen_productos[producto] = resumen_productos.get(producto, 0) + cantidad
                total_valorizado_alimento += buscar_precio_producto(producto, precios) * cantidad

    c1, c2, c3, c4, c5 = st.columns(5)
    c1.markdown(f'<div class="metric-card"><p>Solicitudes</p><div class="big-number">{total_solicitudes}</div></div>', unsafe_allow_html=True)
    c2.markdown(f'<div class="metric-card"><p>Pendientes</p><div class="big-number">{pendientes}</div></div>', unsafe_allow_html=True)
    c3.markdown(f'<div class="metric-card"><p>Entregadas</p><div class="big-number">{entregadas}</div></div>', unsafe_allow_html=True)
    c4.markdown(f'<div class="metric-card"><p>Sacos</p><div class="big-number">{total_sacos}</div></div>', unsafe_allow_html=True)
    c5.markdown(f'<div class="metric-card"><p>Anticipo</p><div class="big-number">{formato_pesos(total_monto)}</div></div>', unsafe_allow_html=True)

    st.write("### 📋 Solicitudes registradas")
    columnas = [
        "id", "fecha", "rut", "nombre", "monto_solicitado",
        "producto1", "cantidad1", "producto2", "cantidad2", "producto3", "cantidad3", "estado"
    ]
    st.dataframe(df_filtrado[[c for c in columnas if c in df_filtrado.columns]], use_container_width=True)

    st.write("### ✅ Entrega por solicitud")
    pendientes_df = df_filtrado[df_filtrado["estado"] == "Pendiente"]

    if pendientes_df.empty:
        st.info("No hay solicitudes pendientes con los filtros seleccionados.")
    else:
        for _, row in pendientes_df.iterrows():
            st.markdown('<div class="row-card">', unsafe_allow_html=True)
            cols = st.columns([1, 2, 3, 3, 2, 2])
            cols[0].write(f"**ID {row['id']}**")
            cols[1].write(f"**RUT:** {row['rut']}")
            cols[2].write(f"**Nombre:** {row['nombre']}")
            detalle = []
            for i in [1, 2, 3]:
                producto = row.get(f"producto{i}")
                cantidad = row.get(f"cantidad{i}")
                if producto and pd.notna(cantidad):
                    detalle.append(f"{producto} ({int(cantidad)})")
            cols[3].write(" / ".join(detalle))
            cols[4].write(f"**Anticipo:** {formato_pesos(row.get('monto_solicitado', 0))}")
            if cols[5].button("✅ Entregar", key=f"entregar_{row['id']}"):
                supabase.table("solicitudes_alimento").update({"estado": "Entregado"}).eq("id", int(row["id"])).execute()
                st.success(f"Solicitud ID {row['id']} marcada como Entregado.")
                st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)

    st.write("### 🔎 Revisión de precios antes de descargar Excel")

    revision_precios = []
    for _, fila in df_filtrado.iterrows():
        for i in [1, 2, 3]:
            producto = fila.get(f"producto{i}")
            cantidad = fila.get(f"cantidad{i}")
            if producto and pd.notna(cantidad):
                precio = buscar_precio_producto(producto, precios)
                revision_precios.append({
                    "Trabajador": fila.get("nombre", ""),
                    "Producto": producto,
                    "Cantidad": int(cantidad),
                    "Precio encontrado": precio,
                    "Total": int(cantidad) * precio,
                    "Revisión": "OK" if precio > 0 else "SIN PRECIO"
                })

    df_revision_precios = pd.DataFrame(revision_precios)

    if not df_revision_precios.empty:
        st.dataframe(df_revision_precios, use_container_width=True)

        sin_precio = df_revision_precios[df_revision_precios["Precio encontrado"] == 0]

        if not sin_precio.empty:
            st.warning("Hay productos sin precio. Revise Administración de productos o cargue los precios oficiales.")
            st.dataframe(sin_precio[["Producto"]].drop_duplicates(), use_container_width=True)
        else:
            st.success("Todos los productos tienen precio asignado.")

    st.write("### 📥 Excel profesional quincenal")
    st.download_button(
        label="📥 Descargar Excel alimento + anticipo",
        data=convertir_excel_formato_quincenal(df_filtrado, mes_seleccionado),
        file_name=f"alimento_y_anticipo_{mes_seleccionado}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

    st.write("### 📦 Total por producto para bodega")
    if resumen_productos:
        df_resumen = pd.DataFrame(list(resumen_productos.items()), columns=["Producto", "Total sacos solicitados"])
        df_resumen["Precio"] = df_resumen["Producto"].apply(lambda p: buscar_precio_producto(p, precios))
        df_resumen["Total valorizado"] = df_resumen["Total sacos solicitados"] * df_resumen["Precio"]
        st.dataframe(df_resumen, use_container_width=True)
        st.download_button(
            "📥 Descargar total productos bodega",
            data=convertir_excel(df_resumen, "Total productos"),
            file_name=f"total_productos_bodega_{mes_seleccionado}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
        st.bar_chart(df_resumen.set_index("Producto")["Total sacos solicitados"])
    else:
        st.info("No hay productos para resumir.")

    st.write("### 💵 Resumen mensual")
    resumen_mensual = pd.DataFrame({
        "Concepto": ["Total sacos", "Total valorizado alimento", "Total anticipo", "Total general"],
        "Valor": [total_sacos, total_valorizado_alimento, total_monto, total_valorizado_alimento + total_monto]
    })
    st.dataframe(resumen_mensual, use_container_width=True)


# =========================================================
# MÓDULO: TRABAJADORES
# =========================================================

if menu == "Trabajadores":
    st.markdown("""
    <div class="header">
        <h1>👷 Administración de trabajadores</h1>
        <h3>Agregar, editar, importar y consultar trabajadores</h3>
    </div>
    """, unsafe_allow_html=True)

    df = obtener_trabajadores_df()

    tab1, tab2, tab3, tab4 = st.tabs(["👤 Ficha", "📂 Importar Excel", "📋 Listado", "📝 Historial"])

    with tab1:
        modo = st.radio("Acción", ["Agregar", "Editar", "Eliminar"], horizontal=True)

        if modo == "Agregar":
            c1, c2 = st.columns(2)
            rut = c1.text_input("RUT")
            nombre = c1.text_input("Nombre completo")
            area = c2.text_input("Área")
            cargo = c2.text_input("Cargo")
            activo = c2.checkbox("Activo", value=True, key="trab_agregar_activo")

            if st.button("✅ Guardar trabajador", use_container_width=True):
                ok, msg = agregar_trabajador({"rut": rut, "nombre": nombre, "area": area, "cargo": cargo, "activo": activo})
                st.success(msg) if ok else st.error(msg)
                if ok:
                    st.rerun()

        elif modo == "Editar":
            if df.empty:
                st.info("No hay trabajadores.")
            else:
                buscar = st.text_input("Buscar trabajador")
                df_b = df.copy()
                if buscar:
                    texto = buscar.lower()
                    df_b = df_b[
                        df_b["rut"].astype(str).str.lower().str.contains(texto, na=False) |
                        df_b["nombre"].astype(str).str.lower().str.contains(texto, na=False)
                    ]

                if not df_b.empty:
                    opciones = [f"{r['nombre']} | {r['rut']}" for _, r in df_b.iterrows()]
                    sel = st.selectbox("Seleccione trabajador", opciones)
                    rut_original = sel.split("|")[-1].strip()
                    trabajador = df[df["rut"] == rut_original].iloc[0]

                    c1, c2 = st.columns(2)
                    rut = c1.text_input("RUT", value=limpiar_texto(trabajador.get("rut")))
                    nombre = c1.text_input("Nombre", value=limpiar_texto(trabajador.get("nombre")))
                    area = c2.text_input("Área", value=limpiar_texto(trabajador.get("area")))
                    cargo = c2.text_input("Cargo", value=limpiar_texto(trabajador.get("cargo")))
                    activo = c2.checkbox("Activo", value=bool(trabajador.get("activo", True)), key=f"trab_editar_activo_{rut_original}")

                    if st.button("💾 Actualizar trabajador", use_container_width=True):
                        ok, msg = actualizar_trabajador(rut_original, {"rut": rut, "nombre": nombre, "area": area, "cargo": cargo, "activo": activo})
                        st.success(msg) if ok else st.error(msg)
                        if ok:
                            st.rerun()

        elif modo == "Eliminar":
            if df.empty:
                st.info("No hay trabajadores.")
            else:
                opciones = [f"{r['nombre']} | {r['rut']}" for _, r in df.iterrows()]
                sel = st.selectbox("Seleccione trabajador", opciones)
                rut_eliminar = sel.split("|")[-1].strip()
                trabajador = df[df["rut"] == rut_eliminar].iloc[0]
                st.warning(f"Eliminará a {trabajador.get('nombre')} - {rut_eliminar}")
                confirmar = st.checkbox("Confirmo eliminación", key="trab_confirmar_eliminar")
                if st.button("🗑️ Eliminar trabajador", use_container_width=True):
                    if confirmar:
                        ok, msg = eliminar_trabajador(rut_eliminar, trabajador.get("nombre", ""))
                        st.success(msg) if ok else st.error(msg)
                        if ok:
                            st.rerun()
                    else:
                        st.error("Debe confirmar.")

    with tab2:
        plantilla = pd.DataFrame({"rut": ["12345678-9"], "nombre": ["JUAN PEREZ"], "area": ["PRODUCCIÓN"], "cargo": ["OPERARIO"], "activo": [True]})
        st.download_button("📥 Descargar plantilla", data=convertir_excel(plantilla, "Plantilla"), file_name="plantilla_trabajadores.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        archivo = st.file_uploader("Subir Excel", type=["xlsx", "xls"])
        if archivo and st.button("📂 Importar trabajadores", use_container_width=True):
            ok, msg, errores = importar_trabajadores_desde_excel(archivo)
            st.success(msg) if ok else st.error(msg)
            if not errores.empty:
                st.dataframe(errores, use_container_width=True)

    with tab3:
        if df.empty:
            st.info("No hay trabajadores.")
        else:
            buscar = st.text_input("Buscar por RUT, nombre, área o cargo")
            filtro = st.selectbox("Estado", ["Todos", "Activos", "Inactivos"], key="trab_filtro_estado")
            df_l = df.copy()

            if buscar:
                texto = buscar.lower()
                condicion = False
                for col in ["rut", "nombre", "area", "cargo"]:
                    if col in df_l.columns:
                        condicion = condicion | df_l[col].astype(str).str.lower().str.contains(texto, na=False)
                df_l = df_l[condicion]

            if filtro == "Activos":
                df_l = df_l[df_l["activo"] == True]
            elif filtro == "Inactivos":
                df_l = df_l[df_l["activo"] == False]

            cols = [c for c in ["rut", "nombre", "area", "cargo", "activo"] if c in df_l.columns]
            st.dataframe(df_l[cols], use_container_width=True)
            st.download_button("📥 Descargar trabajadores", data=convertir_excel(df_l[cols], "Trabajadores"), file_name="trabajadores.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    with tab4:
        hist = cargar_historial_df()
        if hist.empty:
            st.info("No hay historial registrado.")
        else:
            st.dataframe(hist, use_container_width=True)


# =========================================================
# MÓDULO: PRODUCTOS
# =========================================================

if menu == "Productos":
    st.markdown("""
    <div class="header">
        <h1>📦 Administración de productos</h1>
        <h3>Productos, precios y disponibilidad</h3>
    </div>
    """, unsafe_allow_html=True)

    df = obtener_productos_df()

    st.write("### 🔄 Cargar precios oficiales")
    st.info("Aunque Supabase tenga algún precio en 0, el Excel usará los precios oficiales mientras el nombre del producto coincida.")
    if st.button("🔄 Cargar / actualizar precios oficiales", use_container_width=True):
        creados, actualizados, errores = cargar_precios_default_supabase()
        st.success(f"Precios actualizados. Creados: {creados}. Actualizados: {actualizados}.")
        if errores:
            st.dataframe(pd.DataFrame({"Errores": errores}), use_container_width=True)
        st.rerun()

    tab1, tab2, tab3 = st.tabs(["➕ Agregar", "✏️ Editar / eliminar", "📋 Listado"])

    with tab1:
        producto = st.text_input("Producto")
        precio = st.number_input("Precio", min_value=0, step=1000)
        activo = st.checkbox("Activo", value=True, key="prod_agregar_activo")
        if st.button("✅ Guardar producto", use_container_width=True):
            ok, msg = agregar_producto_alimento(producto, precio, activo)
            st.success(msg) if ok else st.error(msg)
            if ok:
                st.rerun()

    with tab2:
        if df.empty:
            st.info("No hay productos.")
        else:
            buscar = st.text_input("Buscar producto")
            df_b = df.copy()
            if buscar:
                df_b = df_b[df_b["producto"].astype(str).str.lower().str.contains(buscar.lower(), na=False)]
            if not df_b.empty:
                opciones = {f"{r['producto']} | ID {r['id']}": r["id"] for _, r in df_b.iterrows()}
                sel = st.selectbox("Seleccione producto", list(opciones.keys()))
                id_producto = opciones[sel]
                actual = df[df["id"] == id_producto].iloc[0]

                producto = st.text_input("Producto", value=limpiar_texto(actual.get("producto")))
                precio = st.number_input("Precio", min_value=0, step=1000, value=int(actual.get("precio", 0) or 0))
                activo = st.checkbox("Activo", value=bool(actual.get("activo", True)), key=f"prod_editar_activo_{id_producto}")

                c1, c2 = st.columns(2)
                if c1.button("💾 Actualizar", use_container_width=True):
                    ok, msg = actualizar_producto_alimento(id_producto, producto, precio, activo)
                    st.success(msg) if ok else st.error(msg)
                    if ok:
                        st.rerun()

                confirmar = c2.checkbox("Confirmo eliminar", key=f"prod_confirmar_eliminar_{id_producto}")
                if c2.button("🗑️ Eliminar", use_container_width=True):
                    if confirmar:
                        ok, msg = eliminar_producto_alimento(id_producto)
                        st.success(msg) if ok else st.error(msg)
                        if ok:
                            st.rerun()
                    else:
                        st.warning("Debe confirmar.")

    with tab3:
        if df.empty:
            st.info("No hay productos.")
        else:
            buscar = st.text_input("Buscar en listado")
            filtro = st.selectbox("Estado", ["Todos", "Activos", "Inactivos"], key="prod_filtro_estado")
            df_l = df.copy()
            if buscar:
                df_l = df_l[df_l["producto"].astype(str).str.lower().str.contains(buscar.lower(), na=False)]
            if filtro == "Activos":
                df_l = df_l[df_l["activo"] == True]
            elif filtro == "Inactivos":
                df_l = df_l[df_l["activo"] == False]

            cols = [c for c in ["id", "producto", "precio", "activo"] if c in df_l.columns]
            st.dataframe(df_l[cols], use_container_width=True)
            st.download_button("📥 Descargar productos", data=convertir_excel(df_l[cols], "Productos"), file_name="productos_alimento.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# =========================================================
# MÓDULO: USUARIOS ADMIN
# =========================================================

if menu == "Usuarios admin":
    st.markdown("""
    <div class="header">
        <h1>👤 Usuarios administradores</h1>
        <h3>Usuarios autorizados para ver administración</h3>
    </div>
    """, unsafe_allow_html=True)

    st.warning("Si todavía no creó la tabla en Supabase, vaya a Configuración y copie el SQL actualizado.")

    tab1, tab2, tab3 = st.tabs(["➕ Agregar usuario", "✏️ Editar usuario", "📋 Listado"])

    with tab1:
        c1, c2 = st.columns(2)
        nuevo_usuario = c1.text_input("Usuario", placeholder="Ej: franco")
        nombre = c1.text_input("Nombre", placeholder="Ej: Franco Gallardo")
        nueva_clave = c2.text_input("Clave", type="password")
        repetir_clave = c2.text_input("Repetir clave", type="password")
        activo = c2.checkbox("Activo", value=True, key="admin_nuevo_activo")

        if st.button("✅ Crear usuario administrador", use_container_width=True):
            if nueva_clave != repetir_clave:
                st.error("Las claves no coinciden.")
            else:
                ok, msg = agregar_usuario_admin(nuevo_usuario, nombre, nueva_clave, activo)
                st.success(msg) if ok else st.error(msg)
                if ok:
                    st.rerun()

    df_admin = obtener_usuarios_admin_df()

    with tab2:
        if df_admin.empty:
            st.info("No hay usuarios administradores guardados en Supabase o falta crear la tabla usuarios_admin.")
        else:
            opciones = {f"{r.get('usuario', '')} | {r.get('nombre', '')} | ID {r.get('id')}": r.get("id") for _, r in df_admin.iterrows()}
            sel = st.selectbox("Seleccione usuario", list(opciones.keys()))
            id_usuario = opciones[sel]
            actual = df_admin[df_admin["id"] == id_usuario].iloc[0]

            c1, c2 = st.columns(2)
            c1.text_input("Usuario", value=limpiar_texto(actual.get("usuario")), disabled=True)
            nombre_editar = c1.text_input("Nombre", value=limpiar_texto(actual.get("nombre")))
            nueva_clave_editar = c2.text_input("Nueva clave (dejar vacío para mantener)", type="password")
            activo_editar = c2.checkbox("Activo", value=bool(actual.get("activo", True)), key=f"admin_activo_{id_usuario}")

            col_a, col_b = st.columns(2)
            if col_a.button("💾 Actualizar usuario", use_container_width=True):
                ok, msg = actualizar_usuario_admin(id_usuario, nombre_editar, nueva_clave_editar or None, activo_editar)
                st.success(msg) if ok else st.error(msg)
                if ok:
                    st.rerun()

            confirmar = col_b.checkbox("Confirmo eliminar", key=f"confirmar_admin_{id_usuario}")
            if col_b.button("🗑️ Eliminar usuario", use_container_width=True):
                if confirmar:
                    ok, msg = eliminar_usuario_admin(id_usuario)
                    st.success(msg) if ok else st.error(msg)
                    if ok:
                        st.rerun()
                else:
                    st.warning("Debe confirmar la eliminación.")

    with tab3:
        if df_admin.empty:
            st.info("No hay usuarios administradores para mostrar.")
        else:
            cols = [c for c in ["id", "usuario", "nombre", "activo", "creado"] if c in df_admin.columns]
            st.dataframe(df_admin[cols], use_container_width=True)
            st.download_button(
                "📥 Descargar usuarios admin",
                data=convertir_excel(df_admin[cols], "Usuarios admin"),
                file_name="usuarios_admin.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )


# =========================================================
# MÓDULO: REPORTES
# =========================================================

if menu == "Reportes":
    st.markdown("""
    <div class="header">
        <h1>📊 Reportes</h1>
        <h3>Resumen general de solicitudes</h3>
    </div>
    """, unsafe_allow_html=True)

    df = preparar_solicitudes_con_nombre()
    if df.empty:
        st.info("No hay datos.")
        st.stop()

    meses = sorted(df["mes"].dropna().unique(), reverse=True)
    mes = st.selectbox("Mes", meses)
    df = df[df["mes"] == mes]

    precios = obtener_precios_productos()
    filas = []
    for _, row in df.iterrows():
        for i in [1, 2, 3]:
            producto = row.get(f"producto{i}")
            cantidad = row.get(f"cantidad{i}")
            if producto and pd.notna(cantidad):
                precio = buscar_precio_producto(producto, precios)
                filas.append({
                    "Producto": producto,
                    "Cantidad": int(cantidad),
                    "Precio": precio,
                    "Total": int(cantidad) * precio
                })

    if filas:
        detalle = pd.DataFrame(filas)
        resumen = detalle.groupby("Producto", as_index=False).agg({"Cantidad": "sum", "Total": "sum"})
        st.write("### Resumen por producto")
        st.dataframe(resumen, use_container_width=True)
        st.bar_chart(resumen.set_index("Producto")["Cantidad"])

    st.write("### Anticipo por trabajador")
    anticipo = df[["rut", "nombre", "monto_solicitado", "estado"]].copy()
    st.dataframe(anticipo, use_container_width=True)



# =========================================================
# MÓDULO: RESET / LIMPIEZA
# =========================================================

if menu == "Reset / limpieza":
    st.markdown("""
    <div class="header">
        <h1>🧹 Reset / limpieza de pruebas</h1>
        <h3>Eliminar registros de prueba antes de usar el portal en la empresa</h3>
    </div>
    """, unsafe_allow_html=True)

    st.error("Zona delicada: estas acciones eliminan datos de Supabase y no se pueden deshacer.")
    st.info("Recomendado antes de partir oficialmente: eliminar solo las solicitudes de prueba. Mantenga trabajadores y productos si ya están correctos.")

    tab1, tab2, tab3 = st.tabs(["🗑️ Eliminar solicitud", "🔄 Reset solicitudes", "⚠️ Reset avanzado"])

    with tab1:
        st.write("### Eliminar una solicitud específica")
        df_sol = preparar_solicitudes_con_nombre()
        if df_sol.empty:
            st.info("No hay solicitudes registradas.")
        else:
            columnas = [c for c in ["id", "fecha", "rut", "nombre", "monto_solicitado", "producto1", "cantidad1", "estado"] if c in df_sol.columns]
            st.dataframe(df_sol[columnas], use_container_width=True)
            id_eliminar = st.number_input("ID de solicitud a eliminar", min_value=1, step=1)
            confirmar_id = st.checkbox("Confirmo que quiero eliminar esta solicitud", key="confirmar_eliminar_solicitud_id")
            if st.button("🗑️ Eliminar solicitud seleccionada", use_container_width=True):
                if not confirmar_id:
                    st.warning("Debe marcar la confirmación antes de eliminar.")
                else:
                    ok, msg = eliminar_solicitud_por_id(id_eliminar)
                    st.success(msg) if ok else st.error(msg)
                    if ok:
                        st.rerun()

    with tab2:
        st.write("### Resetear solicitudes de alimento")
        st.warning("Esto elimina TODAS las solicitudes registradas. No elimina trabajadores, productos ni usuarios administradores.")
        confirmacion = st.text_input("Para confirmar escriba exactamente: ELIMINAR SOLICITUDES", key="txt_reset_solicitudes")
        if st.button("🧹 Eliminar todas las solicitudes", use_container_width=True):
            if confirmacion != "ELIMINAR SOLICITUDES":
                st.error("Texto de confirmación incorrecto.")
            else:
                ok, msg = eliminar_todos_registros_tabla("solicitudes_alimento")
                st.success(msg) if ok else st.error(msg)
                if ok:
                    st.rerun()

        st.write("### Volver solicitudes a pendiente")
        st.info("Esta opción no elimina datos; solo cambia el estado de todas las solicitudes a Pendiente.")
        if st.button("↩️ Marcar todas como Pendiente", use_container_width=True):
            ok, msg = marcar_solicitudes_pendientes()
            st.success(msg) if ok else st.error(msg)
            if ok:
                st.rerun()

    with tab3:
        st.write("### Reset avanzado")
        st.warning("Use esto solo si quiere dejar el portal limpio para comenzar de cero.")
        st.markdown("""
        **Opciones disponibles:**
        - Solicitudes de alimento: borra solicitudes de prueba.
        - Historial trabajadores: borra historial de cambios.
        - Trabajadores: borra la lista de trabajadores.
        - Productos: borra productos y precios cargados.
        
        **No se eliminan los usuarios administradores**, para que no pierda el acceso.
        """)

        borrar_solicitudes = st.checkbox("Eliminar solicitudes de alimento", value=True)
        borrar_historial = st.checkbox("Eliminar historial de trabajadores", value=False)
        borrar_trabajadores = st.checkbox("Eliminar trabajadores", value=False)
        borrar_productos = st.checkbox("Eliminar productos", value=False)

        confirmacion_total = st.text_input("Para confirmar escriba exactamente: RESET PORTAL", key="txt_reset_portal")

        if st.button("⚠️ Ejecutar reset avanzado", use_container_width=True):
            if confirmacion_total != "RESET PORTAL":
                st.error("Texto de confirmación incorrecto.")
            else:
                acciones = []
                if borrar_solicitudes:
                    acciones.append(("solicitudes_alimento", "Solicitudes"))
                if borrar_historial:
                    acciones.append(("historial_trabajadores", "Historial"))
                if borrar_trabajadores:
                    acciones.append(("trabajadores", "Trabajadores"))
                if borrar_productos:
                    acciones.append(("productos_alimento", "Productos"))

                if not acciones:
                    st.warning("Debe seleccionar al menos una opción.")
                else:
                    resultados = []
                    for tabla, nombre in acciones:
                        ok, msg = eliminar_todos_registros_tabla(tabla)
                        resultados.append({"Sección": nombre, "Resultado": "OK" if ok else "ERROR", "Detalle": msg})
                    st.dataframe(pd.DataFrame(resultados), use_container_width=True)
                    if all(r["Resultado"] == "OK" for r in resultados):
                        st.success("Reset ejecutado correctamente.")
                    else:
                        st.warning("Algunas acciones no se pudieron completar. Revise el detalle.")


# =========================================================
# MÓDULO: CONFIGURACIÓN
# =========================================================

if menu == "Configuración":
    st.markdown("""
    <div class="header">
        <h1>⚙️ Configuración del sistema</h1>
        <h3>Parámetros generales, estado y mantenimiento del Portal</h3>
    </div>
    """, unsafe_allow_html=True)

    st.success("El Portal está configurado para que el trabajador ingrese directo a Solicitud trabajador y el administrador entre con usuario y clave.")

    st.write("### 📌 Parámetros actuales")
    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown(f'<div class="metric-card"><p>Día cierre solicitudes</p><div class="big-number">{DIA_CIERRE_SOLICITUD}</div></div>', unsafe_allow_html=True)
    with col2:
        st.markdown(f'<div class="metric-card"><p>Día retiro alimento</p><div class="big-number">{DIA_RETIRO}</div></div>', unsafe_allow_html=True)
    with col3:
        st.markdown(f'<div class="metric-card"><p>Máximo anticipo</p><div class="big-number">{formato_pesos(MONTO_MAXIMO_SOLICITUD)}</div></div>', unsafe_allow_html=True)

    st.write("### 🔐 Accesos")
    st.info("El enlace normal abre solo la solicitud del trabajador. Para administración debes usar el enlace con ?admin=1 al final.")

    enlace_trabajador = "https://alfa-control-app-mvfc9nhfpzwgbjt2dnkfgn.streamlit.app"
    enlace_admin = "https://alfa-control-app-mvfc9nhfpzwgbjt2dnkfgn.streamlit.app/?admin=1"

    c1, c2 = st.columns(2)
    with c1:
        st.write("**Enlace trabajador / QR**")
        st.code(enlace_trabajador)
    with c2:
        st.write("**Enlace administrador**")
        st.code(enlace_admin)

    st.write("### 🧾 Estado de la base de datos")
    try:
        total_trabajadores = len(obtener_trabajadores_df())
        total_productos = len(obtener_productos_df())
        total_solicitudes = len(obtener_solicitudes_df())
        try:
            total_admin = len(obtener_usuarios_admin_df())
        except Exception:
            total_admin = 0

        e1, e2, e3, e4 = st.columns(4)
        e1.metric("Trabajadores", total_trabajadores)
        e2.metric("Productos", total_productos)
        e3.metric("Solicitudes", total_solicitudes)
        e4.metric("Usuarios admin", total_admin)
        st.success("Conexión con Supabase funcionando correctamente.")
    except Exception as e:
        st.error(f"No se pudo verificar Supabase: {e}")

    st.write("### 🧹 Mantenimiento")
    st.warning("Para borrar pruebas o corregir equivocaciones, usa el módulo 'Reset / limpieza'.")

    if st.button("Ir a Reset / limpieza", use_container_width=True):
        st.info("Selecciona 'Reset / limpieza' en el menú lateral para eliminar solicitudes de prueba o registros incorrectos.")

    st.write("### ✅ Recomendación antes de usar en la empresa")
    st.markdown("""
    - Hacer una prueba desde un celular con datos móviles.
    - Crear una solicitud de prueba.
    - Revisarla en Administrador.
    - Descargar el Excel.
    - Luego borrar la prueba desde **Reset / limpieza**.
    """)
